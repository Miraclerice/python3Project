# -*- coding: UTF-8 -*-
# @author  : M_Xie
# @date    : 2022-10-10 21:56:51
import openpyxl

wb = openpyxl.load_workbook("example.xlsx")
# 获取所有sheet名称列表
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
# 根据sheet名称获取对象
sheet1 = wb["北京"]
print(sheet1)

# 获取单元格数据
cell1 = sheet1.cell(1, 1)
print(cell1.value)
# ================================
#  获取所有sheet对象列表
worksheets = wb.worksheets
print(worksheets)
# 获取某一个sheet对象
sheet = worksheets[2]
print(sheet)
cell = sheet.cell(1, 1)
print(cell.value)

# f = open("example.xlsx", "rb")
# wb1 = openpyxl.load_workbook(f)
# print(wb1)
# for sheet in wb1:
#     print(sheet.title)
# f.close()
